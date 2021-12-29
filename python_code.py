import os
from pathlib import Path
import openpyxl

def to_delete(filename,ori_col,ori_row,path_to_folder):
    initial_path = Path(path_to_folder)
    path_file_list = os.listdir(path_to_folder)
    wb = openpyxl.load_workbook(filename)
    sheet_interest = wb.sheetnames[0]  # select first sheet
    sheet = wb[sheet_interest]
    max_row = sheet.max_row
    files_to_delete = []
    for i in range(0, max_row - ori_row + 1):
        files_to_delete = files_to_delete + [sheet.cell(row=ori_row + i, column=ori_col).value]
    for k in range(len(files_to_delete)):
        if files_to_delete[k] in path_file_list:
            final_path = os.path.join(initial_path, files_to_delete[k])
            send2trash.send2trash(final_path)
        else:
            print(files_to_delete[k],'is not available in',initial_path)
